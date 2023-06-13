# coding:utf-8
# Name:read_excel.py
# Author:nigo
# Time:2021/7/7 5:16 下午
# Description:
from openpyxl import load_workbook


class Doexcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []  # 数据存储为列表格式
        for i in range(1, sheet.max_row):
            sub_data = {}  # 数据存储为字典格式
            sub_data['url'] = sheet.cell(i + 1, 1).value
            sub_data['data'] = sheet.cell(i + 1, 2).value
            sub_data['method'] = sheet.cell(i + 1, 3).value
            print(sub_data)
            test_data.append(sub_data)

        return test_data  # 返回获取到的数据
    # def get_data1(self):
    #     wb = load_workbook(self.file_name)
    #     sheet = wb[self.sheet_name]
    #     test_data = []
    #     for i in range(1,sheet.max_row):
    #         sub_data = {}
    #         sub_data['url'] = sheet.cell(i+1,1).value
    #         sub_data['data'] = sheet.cell(i+1,2).value
    #         test_data.append(sub_data)
    #
    #     return test_data


print(Doexcel(r'../test_data/Doexcel.xlsx', 'Sheet1').get_data())

